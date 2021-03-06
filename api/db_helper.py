import sqlite3
from sqlite3 import Error, Connection
from openpyxl import Workbook, load_workbook, styles
from os import path

def get_sqlconnection():
    """
    Get sql connection for sqlite
    """
    con: Connection = None
    try:
        con = sqlite3.connect('dbmonitor.db')
        con.row_factory = sqlite3.Row

        print("Connection is established: Database is created.")
        return con
    except Error as DbError:
        print(DbError)


def create_table(con: Connection) -> Connection:
    """
    Create a database to store the databases to be monitored
    """
    db_cursor = con.cursor()
    sql_script = """ CREATE TABLE IF NOT EXISTS tb_monitor
        (id text PRIMARY KEY,
         db_name text,
         size float, 
         monitor_time  date)"""
    db_cursor.execute(sql_script)
    con.commit()


def insert_record(con: Connection, entries: tuple):
    """
    Insert into the monitored tables
    """
    cursorObj = con.cursor()
    cursorObj.executemany('''INSERT INTO tb_monitor
        (id, db_name, size, monitor_time) 
        VALUES(?, ?, ?, ?)''', entries)

    con.commit()


def get_all_records(con: Connection):
    """
    Get a list of records of the monitored databases
    """
    try:
        cursorObj = con.cursor()
        cursorObj.execute(
            'SELECT id, db_name, size, monitor_time FROM tb_monitor GROUP BY db_name ORDER BY monitor_time DESC')
        rows = cursorObj.fetchall()
        rowarray_list = []
        for row in rows:
            d = dict(zip(row.keys(), row))   # a dict with column names as keys
            rowarray_list.append(d)

        return rowarray_list
    except Error as DbError:
        print(DbError)
    # finally:
        # close_connection(con)


def get_records_between_date_range(con: Connection, date_range: dict):
    """
    Get a list of records of the monitored databases
    """
    try:
        cursorObj = con.cursor()
        cursorObj.execute('SELECT * FROM tb_monitor WHERE monitor_time BETWEEN ? AND ?',
                          date_range.start, date_range.end)
        rows = cursorObj.fetchall()
        return rows
    except Error as DbError:
        print(DbError)


def close_connection(con: Connection):
    """
    Close connection
    """
    con.close()


def add_record_to_excel(data):
    """
    Create a new record to excel file or add to existing file
    """ 
    EXCEL_FILE= "dbmonitor.xlsx"
 
    file_exists = path.exists(EXCEL_FILE)
    if file_exists:
        book = load_workbook(EXCEL_FILE)
        ws = book.active

        for item in data:
            ws.append(list(item.values()))

        
    else:
        print("new file")
        book = Workbook()
        ws = book.active

        dbs = []
        header = list(data[0].keys()) # get the  headers from the keys
        ws.append(header)

        row = ws.row_dimensions[1]
        row.font = styles.Font(bold=True, size=13)

        for item in data:
            ws.append(list(item.values()))    

    book.save("dbmonitor.xlsx")